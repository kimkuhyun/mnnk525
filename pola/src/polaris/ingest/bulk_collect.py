"""scripts/bulk_collect.py — 자족 단일 파일.

카탈로그(docs/APIdocs/API_메타카탈로그.xlsx) 기반 5사 일괄 수집기.
APIpipe_미완성 패키지 없이 동작 (polaris 모듈을 본 파일에 인라인).

수집 단계:
  1. DART OpenAPI (DS001~DS005, callable 엔드포인트 전부)
  2. 사업보고서 원문 (PDF + HTML)  — DART 뷰어, rate-limit 보수적
  3. KRX 일별 OHLCV  — FinanceDataReader (API 키 불필요)
  4. 뉴스 (knews-rss 25피드, since 필터 + 회사 매칭 메타 부착)
  5. FTC 공정위 OpenAPI (대규모기업집단 + 5사 관련 그룹 소속회사·업종)

기본 5사: 삼성전자·SK하이닉스·동진쎄미켐·솔브레인(사업회사)·한미반도체

산출물: ___test/rawData/{corp_code}/{dart|documents|krx|news}/ + _common/

사용법:
  python scripts/bulk_collect.py                        # 기본 5사 × 25년~현재
  python scripts/bulk_collect.py --corps "373220"       # 다른 회사
  python scripts/bulk_collect.py --only krx,news        # 일부 단계만
  python scripts/bulk_collect.py --from-year 2024       # 기간 조정
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import sys
import time
import zipfile
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import unquote
from xml.etree import ElementTree as ET

import feedparser
import httpx
import typer
from dotenv import load_dotenv
from loguru import logger
from lxml import html as lxhtml
from openpyxl import load_workbook
from readability import Document
from tenacity import (
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential,
)

# KRX 시계열 — FinanceDataReader (API 키 불필요)
try:
    import FinanceDataReader as fdr
except ImportError:
    fdr = None

# ====== 경로 + .env 로드 ======
# polaris.config.DATA_ROOT = {PKG_ROOT}/data. PKG_ROOT = 패키지 루트 (pola/).
from polaris.config import DATA_ROOT, ROOT as PKG_ROOT
load_dotenv(PKG_ROOT / ".env", override=False)

DART_API_KEY = os.getenv("DART_API_KEY", "")
FTC_OPENAPI_KEY = os.getenv("FTC_OPENAPI_KEY", "")
BOK_ECOS_KEY = os.getenv("BOK_ECOS_KEY", "")
KOSIS_KEY = os.getenv("KOSIS_KEY", "")

app = typer.Typer(add_completion=False)

# ====== 기본 매핑 ======

# ISIN fallback (corps.json 에는 없음. 새 회사는 KRX API 가 자동 조회).
_ISIN_FALLBACK = {
    "00126380": "KR7005930003", "00164779": "KR7000660001",
    "00118804": "KR7005290002", "01489648": "KR7357780006",
    "00161383": "KR7042700009",
}


def _build_default_corps() -> dict[str, dict[str, str]]:
    """polaris.config.CORPS (.env 단일 source) + corps.json 자동 조회로 빌드.
    새 회사 추가 = .env POLARIS_CORPS 에 corp_code 만 추가하면 자동 적용."""
    from polaris.config import CORPS as _CORPS, get_corp_meta
    out: dict[str, dict[str, str]] = {}
    for cc in _CORPS:
        m = get_corp_meta(cc)
        out[cc] = {
            "stock": m.get("stock_code", ""),
            "name": m.get("corp_name", cc),
            "isin": _ISIN_FALLBACK.get(cc, ""),
        }
    return out


DEFAULT_CORPS: dict[str, dict[str, str]] = _build_default_corps()

DEFAULT_CATALOG = PKG_ROOT / "docs" / "APIdocs" / "API_메타카탈로그.xlsx"
DEFAULT_OUT = DATA_ROOT / "rawData"

REPRT_CODES = ["11013", "11012", "11014", "11011"]  # 1Q·반기·3Q·사업
IDX_CL_CODES = ["M210000", "M220000", "M230000", "M240000"]

# ====== Rate-limit 프로필 ======
# DART OpenAPI: 분당 1000회 한도, 일 20,000회. normal=분당 ~300회 (30%)
# DART 뷰어 (dart.fss.or.kr): 분당 30회+ 연속 호출 시 IP 차단. normal=회사간 30초.
# slow: 30사+ 안전 / normal: 5사 기본 / fast: 개발 디버깅용 (IP 차단 위험)
RATE_PROFILES = {
    "slow": {
        "SLEEP_DART_REQUEST": 0.5, "SLEEP_DART_RATELIMIT": 10.0,
        "SLEEP_PDF_BEFORE": 5.0, "SLEEP_PDF_BETWEEN": 3.0,
        "SLEEP_RNO_GAP": 4.0, "SLEEP_CORP_GAP": 60.0,
        "SLEEP_NEWS_ENTRY": 2.0, "SLEEP_FTC_REQUEST": 1.0,
        "SLEEP_KRX_CORP": 2.0, "PDF_RETRY_BACKOFF": (10.0, 30.0, 60.0),
    },
    "normal": {
        "SLEEP_DART_REQUEST": 0.20, "SLEEP_DART_RATELIMIT": 5.0,
        "SLEEP_PDF_BEFORE": 3.0, "SLEEP_PDF_BETWEEN": 2.0,
        "SLEEP_RNO_GAP": 2.0, "SLEEP_CORP_GAP": 30.0,
        "SLEEP_NEWS_ENTRY": 1.0, "SLEEP_FTC_REQUEST": 0.5,
        "SLEEP_KRX_CORP": 1.0, "PDF_RETRY_BACKOFF": (5.0, 15.0, 30.0),
    },
    "fast": {
        "SLEEP_DART_REQUEST": 0.05, "SLEEP_DART_RATELIMIT": 2.0,
        "SLEEP_PDF_BEFORE": 1.0, "SLEEP_PDF_BETWEEN": 0.5,
        "SLEEP_RNO_GAP": 0.5, "SLEEP_CORP_GAP": 5.0,
        "SLEEP_NEWS_ENTRY": 0.3, "SLEEP_FTC_REQUEST": 0.1,
        "SLEEP_KRX_CORP": 0.3, "PDF_RETRY_BACKOFF": (2.0, 5.0, 10.0),
    },
}

# 기본은 normal — apply_rate_profile()이 main() 시작 시 갱신
SLEEP_DART_REQUEST = 0.20
SLEEP_DART_RATELIMIT = 5.0
SLEEP_PDF_BEFORE = 3.0
SLEEP_PDF_BETWEEN = 2.0
SLEEP_RNO_GAP = 2.0
SLEEP_CORP_GAP = 30.0
SLEEP_NEWS_ENTRY = 1.0
SLEEP_FTC_REQUEST = 0.5
SLEEP_KRX_CORP = 1.0
PDF_RETRY_BACKOFF = (5.0, 15.0, 30.0)

def apply_rate_profile(profile: str) -> None:
    """CLI 옵션 --profile 적용. 모듈 전역 텀 갱신."""
    if profile not in RATE_PROFILES:
        raise ValueError(f"profile은 {list(RATE_PROFILES)} 중 하나. 받음: {profile!r}")
    g = globals()
    for k, v in RATE_PROFILES[profile].items():
        g[k] = v
    logger.info(f"Rate-limit profile: '{profile}' 적용 — CORP_GAP={g['SLEEP_CORP_GAP']}s, RNO_GAP={g['SLEEP_RNO_GAP']}s, DART_REQ={g['SLEEP_DART_REQUEST']}s")

# ====== Endpoint 타입 분류 ======

EP_TYPES: dict[str, str] = {
    "company": "corp_code_only",
    "majorstock": "corp_code_only",
    "elestock": "corp_code_only",
    "list": "list",
    "corpCode": "no_arg_zip",
    "document": "rcept_no",
    "fnlttXbrl": "rcept_no_reprt",
    "xbrlTaxonomy": "sj_div",
    "fnlttMultiAcnt": "multi_corp_year_reprt",
    "fnlttCmpnyIndx": "multi_corp_year_reprt_idx",
    "fnlttSinglIndx": "corp_year_reprt_idx",
    "fnlttSinglAcntAll": "corp_year_reprt_fs",
}

def _ep_type(endpoint: str, category: str) -> str:
    if endpoint in EP_TYPES:
        return EP_TYPES[endpoint]
    if category == "DS005":
        return "corp_period"
    if category in ("DS002", "DS003"):
        return "corp_year_reprt"
    if category == "FTC":
        return "ftc"
    return "skip"

# ====== 보고서 종류 필터 ======
# G2: 코드 상수 + catalog/8.다운로드_문서종류_선별 자동 로드 (SSOT).
# load_report_filters_from_catalog()가 main() 시작 시 catalog/8 시트 결정(KEEP/DROP)
# 기준으로 정확 일치 set + 보조 substring 키워드를 빌드. 코드 상수는 fallback.

# Fallback (catalog/8 로드 실패 시 또는 신규 종류 substring 매칭용)
KEY_REPORT_KEYWORDS = (
    "사업보고서", "반기보고서", "분기보고서",
    "단일판매", "공급계약", "수주",
    "주요사항보고서",
    "유상증자", "무상증자", "유무상증자", "감자",
    "자기주식",
    "회사분할", "회사합병", "회사분할합병",
    "영업양수", "영업양도", "영업정지",
    "타법인", "주식양수도", "주식교환",
    "전환사채", "신주인수권부사채", "교환사채",
    "기업설명회",
    "동일인등출자계열회사",
    "대규모기업집단",
    "주주총회",
    "감사보고서", "감사의견",
    "현금ㆍ현물배당", "현금·현물배당", "현금배당",
    "정정",
    "특수관계인과의", "의결권대리", "잠정실적", "잠정)실적",
    "주식변동신고서", "신규시설투자",
)

DROP_REPORT_PATTERNS = (
    "기업설명회", "임원ㆍ주요주주특정증권", "임원ㆍ주요주주 특정증권",
    "주주총회소집결의", "임시주주총회결과",
    "주주총회집중일", "주주총회 집중일",
    "주주명부폐쇄", "수시공시의무관련사항",
    "특수관계인에대한", "주식등의대량보유상황",
)

# Catalog/8 로드 시 채워지는 정확 일치 set (SSOT 우선)
CATALOG_KEEP_SET: set[str] = set()
CATALOG_DROP_SET: set[str] = set()

def _normalize_report_nm(s: str) -> str:
    """[기재정정] prefix 제거 + 공백 정규화."""
    s = re.sub(r'\s+', ' ', s or '').strip()
    if s.startswith('['):
        c = s.find(']')
        if c > 0:
            return s[c+1:].strip()
    return s

def load_report_filters_from_catalog(catalog_path: Path) -> tuple[int, int]:
    """catalog/8.다운로드_문서종류_선별 시트에서 KEEP/DROP report_nm 추출 → 모듈 전역 set 갱신.
    Returns (keep_n, drop_n)."""
    global CATALOG_KEEP_SET, CATALOG_DROP_SET
    try:
        wb = load_workbook(catalog_path, read_only=True, data_only=True)
        try:
            if '8.다운로드_문서종류_선별' not in wb.sheetnames:
                logger.warning("catalog에 8.다운로드_문서종류_선별 시트 없음 — 코드 상수만 사용")
                return (0, 0)
            ws = wb['8.다운로드_문서종류_선별']
            keep, drop = set(), set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                nm = row[0]
                if not nm or str(nm).startswith('─'):
                    continue
                dec = row[7] if len(row) > 7 else None
                nm_n = _normalize_report_nm(str(nm))
                if dec == 'KEEP':
                    keep.add(nm_n)
                elif dec == 'DROP':
                    drop.add(nm_n)
            CATALOG_KEEP_SET = keep
            CATALOG_DROP_SET = drop
            logger.info(f"catalog/8 로드: KEEP {len(keep)}종 / DROP {len(drop)}종")
            return (len(keep), len(drop))
        finally:
            wb.close()
    except Exception as e:
        logger.warning(f"catalog/8 로드 실패: {e} — 코드 상수만 사용")
        return (0, 0)

def _is_drop_report(report_nm: str) -> bool:
    if not report_nm:
        return False
    nm_n = _normalize_report_nm(report_nm)
    # 1순위: catalog/8 DROP set 정확 일치
    if nm_n in CATALOG_DROP_SET:
        return True
    # 2순위: 코드 상수 substring (fallback)
    return any(p in report_nm for p in DROP_REPORT_PATTERNS)

def _is_key_report(report_nm: str) -> bool:
    """사용자 정책: 사업·반기·분기 보고서 3종만 PDF 다운로드.
    나머지 공시 (자기주식·임원변동·정정 등) 는 DART JSON 으로 충분.
    [기재정정]사업보고서 같은 정정도 substring 매칭으로 포함됨."""
    if not report_nm:
        return False
    return any(t in report_nm for t in ("사업보고서", "반기보고서", "분기보고서"))

# ====== 헬퍼 ======

def short_hash(s: str, n: int = 8) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()[:n]

def params_hash(params: dict) -> str:
    safe = {k: v for k, v in (params or {}).items() if k != "crtfc_key"}
    raw = json.dumps(safe, sort_keys=True, ensure_ascii=False)
    return short_hash(raw)

def news_id_from_url(url: str) -> str:
    """뉴스 자연키 = sha1(url)[:16]."""
    return hashlib.sha1((url or "").strip().encode("utf-8")).hexdigest()[:16]

def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

def yyyymmdd(d: date) -> str:
    return d.strftime("%Y%m%d")

# ====== DART OpenAPI 클라이언트 (polaris.adapters.base 인라인) ======

DART_BASE_URL = "https://opendart.fss.or.kr/api"

# DART status 코드 → 정책
DART_STATUS_MAP: dict[str, str] = {
    "000": "ok", "013": "no_data",
    "010": "key_error", "011": "key_error",
    "020": "rate_limit", "100": "bad_param", "101": "bad_param",
    "800": "maintenance", "900": "unknown_error",
}

@dataclass
class DartResponse:
    endpoint: str
    status: str
    raw_status_code: str
    message: str
    data: Any
    cached: bool = False
    params: dict | None = None

    @property
    def is_ok(self) -> bool:
        return self.status == "ok"

class DartKeyError(RuntimeError):
    """DART API 키 오류 — 즉시 중단."""

class DartRateLimitError(RuntimeError):
    """DART 한도 초과 — 백오프 후 재시도."""

class DartHTTPClient:
    """DART OpenAPI 클라이언트 — on-disk 캐시 + tenacity 재시도 + status 분류."""

    def __init__(self, api_key: str, cache_dir: Path, timeout: float = 30.0):
        if not api_key:
            raise ValueError("DART_API_KEY가 비어있음 (.env 확인)")
        self._api_key = api_key
        self._cache_dir = Path(cache_dir)
        self._cache_dir.mkdir(parents=True, exist_ok=True)
        self._client = httpx.Client(
            base_url=DART_BASE_URL, timeout=timeout, follow_redirects=True,
        )

    def close(self) -> None:
        self._client.close()

    def __enter__(self) -> "DartHTTPClient":
        return self

    def __exit__(self, *a: object) -> None:
        self.close()

    def request(
        self, endpoint: str, params: dict | None = None,
        fmt: str = "json", url_ext: str | None = None,
        fresh: bool = False, cache_subdir: str | None = None,
    ) -> DartResponse:
        params = dict(params or {})
        params["crtfc_key"] = self._api_key
        cache_root = self._cache_dir / (cache_subdir or "_root")
        cache_root.mkdir(parents=True, exist_ok=True)
        ck = params_hash(params)
        cache_file = cache_root / f"{endpoint}_{ck}.{fmt}"

        if cache_file.exists() and not fresh:
            data = self._parse_cached(cache_file, fmt)
            return self._classify(endpoint, data, params, cached=True)

        ext = url_ext or fmt
        url = f"{endpoint}.{ext}"
        try:
            data = self._do_call(url, params, fmt)
        except DartKeyError:
            raise
        except DartRateLimitError:
            raise
        except httpx.HTTPError as e:
            return DartResponse(
                endpoint=endpoint, status="http_error",
                raw_status_code=f"http_{type(e).__name__}",
                message=str(e), data=None, params=params,
            )
        try:
            self._save_cached(cache_file, data, fmt)
        except Exception as e:
            logger.warning(f"cache 저장 실패 {cache_file}: {e}")
        # 보수적 rate-limit
        time.sleep(SLEEP_DART_REQUEST)
        return self._classify(endpoint, data, params, cached=False)

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=2, max=10),
        retry=retry_if_exception_type(
            (httpx.TransportError, httpx.HTTPStatusError, DartRateLimitError)
        ),
        reraise=True,
    )
    def _do_call(self, url: str, params: dict, fmt: str) -> Any:
        r = self._client.get(url, params=params)
        if r.status_code >= 400:
            if r.status_code >= 500:
                r.raise_for_status()
            return {"status": f"http_{r.status_code}", "message": r.text[:200]}
        if fmt == "json":
            try:
                payload = r.json()
            except json.JSONDecodeError as e:
                return {"status": "parse_error", "message": str(e)}
            st = payload.get("status")
            if st == "020":
                time.sleep(SLEEP_DART_RATELIMIT)
                raise DartRateLimitError(payload.get("message", "rate limit"))
            return payload
        elif fmt == "xml":
            return r.text
        elif fmt == "zip":
            return r.content
        return r.text

    def _classify(self, endpoint: str, data: Any, params: dict, cached: bool) -> DartResponse:
        if not isinstance(data, dict):
            return DartResponse(
                endpoint=endpoint, status="ok", raw_status_code="binary",
                message="binary response", data=data, cached=cached, params=params,
            )
        st_raw = str(data.get("status", "unknown"))
        msg = str(data.get("message", ""))
        if st_raw.startswith("http_"):
            status = "http_error"
        elif st_raw == "parse_error":
            status = "parse_error"
        else:
            status = DART_STATUS_MAP.get(st_raw, "unknown_error")
        if status == "key_error":
            logger.error(f"DART 키 오류 {endpoint}: status={st_raw} msg={msg!r}")
            raise DartKeyError(f"DART status={st_raw}: {msg}")
        return DartResponse(
            endpoint=endpoint, status=status, raw_status_code=st_raw,
            message=msg, data=data, cached=cached, params=params,
        )

    @staticmethod
    def _save_cached(path: Path, data: Any, fmt: str) -> None:
        if fmt == "json":
            path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        elif fmt == "xml":
            path.write_text(data, encoding="utf-8")
        elif fmt == "zip":
            path.write_bytes(data)

    @staticmethod
    def _parse_cached(path: Path, fmt: str) -> Any:
        if fmt == "json":
            return json.loads(path.read_text(encoding="utf-8"))
        if fmt == "xml":
            return path.read_text(encoding="utf-8")
        if fmt == "zip":
            return path.read_bytes()
        return path.read_text(encoding="utf-8")

# ====== 사업보고서 PDF/HTML (polaris.adapters.dart_documents 인라인) ======

DART_VIEWER_BASE = "https://dart.fss.or.kr"
_DCMNO_RE = re.compile(r"openPdfDownload\(\s*['\"](\d+)['\"]\s*,\s*['\"](\d+)['\"]")

def extract_html_zip(zip_bytes: bytes, out_dir: Path) -> tuple[Path | None, list[Path]]:
    """ZIP을 풀어 body html + 첨부 저장. (body_path, extras)."""
    out_dir.mkdir(parents=True, exist_ok=True)
    body_path: Path | None = None
    extras: list[Path] = []
    if not zip_bytes.startswith(b"PK"):
        raise RuntimeError(f"ZIP signature 없음 (size={len(zip_bytes)})")
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        for name in z.namelist():
            try:
                data = z.read(name)
            except Exception as e:
                logger.warning(f"ZIP entry {name} 읽기 실패: {e}")
                continue
            safe = name.replace("\\", "/").lstrip("/")
            target = out_dir / safe
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(data)
            lower = safe.lower()
            if (lower.endswith(".xml") or lower.endswith(".html")) and body_path is None:
                body_path = target
            elif lower.endswith((".xml", ".html", ".htm")):
                extras.append(target)
    if body_path is None:
        first = sorted(out_dir.rglob("*.*"))
        if first:
            body_path = first[0]
    return body_path, extras

def find_pdf_url(http: httpx.Client, rcept_no: str) -> tuple[str | None, str | None]:
    """DART 뷰어 → dcm_no → 다운로드 페이지에서 진짜 pdf.do URL 추출.

    Returns (pdf_url, referer_url). PDF가 없는 공시(짧은 공시 등 HTML만)는 (None, None).
    pdf.do는 세션(JSESSIONID) + Referer(main.do) 필수 — 동일 client 안에서 호출되어야 함.
    """
    viewer_url = f"{DART_VIEWER_BASE}/dsaf001/main.do?rcpNo={rcept_no}"
    try:
        # 1) viewer GET — 세션 발급 + dcm_no 추출
        rv = http.get(viewer_url, timeout=20, follow_redirects=True)
        if rv.status_code != 200:
            return None, None
        m = _DCMNO_RE.search(rv.text)
        if not m:
            return None, None
        rno, dcm = m.group(1), m.group(2)
        # 2) main.do GET — 다운로드 페이지 (Referer=viewer)
        main_url = f"{DART_VIEWER_BASE}/pdf/download/main.do?rcp_no={rno}&dcm_no={dcm}"
        rm = http.get(main_url, headers={"Referer": viewer_url}, timeout=20)
        if rm.status_code != 200:
            return None, None
        # 3) main.do 응답에 pdf.do 링크 있나?
        #    짧은 공시(배당결정 등)는 zip.do(HTML)만 있고 pdf.do 링크 없음 → 자동 skip
        if "/pdf/download/pdf.do?" not in rm.text:
            return None, None
        pdf_url = f"{DART_VIEWER_BASE}/pdf/download/pdf.do?rcp_no={rno}&dcm_no={dcm}"
        return pdf_url, main_url
    except Exception as e:
        logger.warning(f"PDF URL 추출 실패 ({rcept_no}): {e}")
        return None, None

def download_pdf(http: httpx.Client, rcept_no: str, pdf_url: str, out_path: Path,
                 referer: str | None = None) -> bool:
    """PDF 다운로드 → out_path. Referer=main.do 필수 (안 주면 빈 응답)."""
    headers = {"Referer": referer} if referer else {}
    try:
        with http.stream("GET", pdf_url, timeout=60, follow_redirects=True,
                         headers=headers) as r:
            if r.status_code != 200:
                return False
            ct = r.headers.get("content-type", "").lower()
            out_path.parent.mkdir(parents=True, exist_ok=True)
            with open(out_path, "wb") as f:
                for chunk in r.iter_bytes():
                    f.write(chunk)
        size = out_path.stat().st_size
        if size < 1000:
            return False
        head = out_path.read_bytes()[:4]
        return head.startswith(b"%PDF") or "pdf" in ct or "octet" in ct
    except Exception as e:
        logger.warning(f"PDF 다운로드 실패 ({rcept_no}): {e}")
        return False

@dataclass
class DocumentResult:
    rcept_no: str
    corp_code: str
    body_html_path: Path | None = None
    pdf_path: Path | None = None
    pdf_url: str | None = None
    error: str | None = None

def fetch_business_report(
    zip_bytes: bytes, rcept_no: str, corp_code: str,
    out_dir: Path, http_pdf: httpx.Client,
) -> DocumentResult:
    """ZIP 풀어 body.html + PDF (재시도)."""
    out_dir.mkdir(parents=True, exist_ok=True)
    result = DocumentResult(rcept_no=rcept_no, corp_code=corp_code)
    (out_dir / "source.zip").write_bytes(zip_bytes)
    try:
        body, _ = extract_html_zip(zip_bytes, out_dir / "extracted")
        if body is not None:
            target = out_dir / "body.html"
            target.write_bytes(body.read_bytes())
            result.body_html_path = target
    except Exception as e:
        result.error = f"ZIP 해제 실패: {e}"
    # PDF — 3회 재시도, 점진 백오프. main.do에 pdf.do 링크 없으면 PDF 없는 공시 → 즉시 skip.
    time.sleep(SLEEP_PDF_BEFORE)
    for attempt, backoff in enumerate(PDF_RETRY_BACKOFF):
        pdf_url, referer = find_pdf_url(http_pdf, rcept_no)
        if pdf_url is None:
            # main.do 응답에 pdf.do 링크 없음 = HTML-only 공시. 재시도 무의미.
            return result
        if pdf_url:
            time.sleep(SLEEP_PDF_BETWEEN)
            pdf_target = out_dir / "original.pdf"
            if download_pdf(http_pdf, rcept_no, pdf_url, pdf_target, referer):
                result.pdf_path = pdf_target
                result.pdf_url = pdf_url
                return result
        time.sleep(backoff)
    return result

# ====== 뉴스 RSS (polaris.adapters.news_rss 인라인) ======

KNEWS_FEEDS_CSV = "https://raw.githubusercontent.com/akngs/knews-rss/main/data/feed_specs.csv"
ALLOWED_CATEGORIES = {"economy", "tech", "science", "_all_"}

# POLARIS 뉴스 정책: 반도체 5사 비즈니스 분석에 최적화된 3 RSS 만.
# (한경 증권 + 매경 기업/경영 + 매경 증권 — 정확도·노이즈 최소)
POLARIS_NEWS_FEEDS: list[dict] = [
    {"feed_id": "hankyung_finance",
     "url": "https://www.hankyung.com/feed/finance",
     "publisher": "한국경제", "title": "한경 증권",
     "categories": "economy"},
    {"feed_id": "mk_business",
     "url": "https://www.mk.co.kr/rss/50100032/",
     "publisher": "매일경제", "title": "매경 기업/경영",
     "categories": "economy"},
    {"feed_id": "mk_securities",
     "url": "https://www.mk.co.kr/rss/50200011/",
     "publisher": "매일경제", "title": "매경 증권",
     "categories": "economy"},
]

@dataclass
class NewsItem:
    news_id: str
    url: str
    title: str
    published: str
    source_feed: str
    source_publisher: str
    text: str

def fetch_feed_specs(http: httpx.Client) -> list[dict]:
    r = http.get(KNEWS_FEEDS_CSV, timeout=30)
    r.raise_for_status()
    reader = csv.DictReader(io.StringIO(r.text))
    return list(reader)

def select_feeds(feeds: list[dict]) -> list[dict]:
    # POLARIS 정책: knews-rss CSV 무시하고 3 RSS 강제 (한경/매경 핵심만).
    return POLARIS_NEWS_FEEDS


# 한경 RSS 등 일부 피드가 HTML named entity(&nbsp; 등) 와 bare & 를 그대로 둬서
# feedparser 의 SAX 파서가 "undefined entity" 로 entries=0 반환.
# raw bytes 받아 정리한 뒤 feedparser 에 넘기면 정상 파싱됨.
_RSS_NAMED_ENT = {
    "nbsp": " ", "copy": "(c)", "reg": "(R)", "trade": "(TM)",
    "hellip": "...", "ndash": "-", "mdash": "--",
    "lsquo": "'", "rsquo": "'", "ldquo": '"', "rdquo": '"',
    "laquo": "<<", "raquo": ">>", "times": "x", "divide": "/",
    "deg": "deg", "euro": "EUR", "pound": "GBP", "yen": "JPY", "cent": "c",
    "middot": "*", "bull": "*",
}
_RSS_NAMED_ENT_RE = re.compile(r"&(" + "|".join(_RSS_NAMED_ENT) + r");")
_RSS_BARE_AMP_RE = re.compile(r"&(?!amp;|lt;|gt;|quot;|apos;|#\d+;|#x[0-9a-fA-F]+;)")


def _sanitize_rss(text: str) -> str:
    text = _RSS_NAMED_ENT_RE.sub(lambda m: _RSS_NAMED_ENT[m.group(1)], text)
    text = _RSS_BARE_AMP_RE.sub("&amp;", text)
    return text


def fetch_feed_parsed(http: httpx.Client, url: str):
    """feed URL → feedparser. sanitize 후 parse 해서 한경 RSS undefined entity 도 처리."""
    try:
        r = http.get(url, timeout=20)
        if r.status_code != 200:
            return feedparser.parse(b"")  # 빈 객체
        return feedparser.parse(_sanitize_rss(r.text))
    except Exception as e:
        logger.warning(f"feed fetch {url}: {e}")
        return feedparser.parse(b"")


# news_raw RDB SSOT 저장 (POLARIS_NEWS_FEEDS 매핑)
_URL_TO_FEED_ID = {f["url"]: f["feed_id"] for f in POLARIS_NEWS_FEEDS}
_FEED_TO_CAT = {f["feed_id"]: f.get("title", "") for f in POLARIS_NEWS_FEEDS}


def _news_raw_insert(item, feed: dict) -> None:
    """news_raw 테이블에 INSERT (idempotent ON DUPLICATE)."""
    from polaris.config import mariadb_conn as _mc
    from email.utils import parsedate_to_datetime
    feed_id = _URL_TO_FEED_ID.get(item.source_feed, "")
    category = _FEED_TO_CAT.get(feed_id, "")
    pub_dt = None
    if item.published:
        try:
            pub_dt = parsedate_to_datetime(item.published).replace(tzinfo=None)
        except Exception:
            pass
    try:
        conn = _mc(); cur = conn.cursor()
        cur.execute("""INSERT INTO news_raw
            (news_id, feed_id, publisher, category, title, url, published, body)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
              title=VALUES(title), url=VALUES(url),
              published=VALUES(published), body=VALUES(body)""",
            (item.news_id, feed_id, item.source_publisher, category,
             item.title[:500], item.url[:1024], pub_dt, item.text))
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        logger.warning(f"news_raw insert {item.news_id}: {e}")

def extract_article_text(html: str) -> str:
    try:
        doc = Document(html)
        tree = lxhtml.fromstring(doc.summary())
        text = " ".join(tree.itertext())
        return re.sub(r"\s+", " ", text).strip()
    except Exception:
        return ""

def fetch_article_body(http: httpx.Client, url: str) -> str:
    try:
        r = http.get(
            url, timeout=15, follow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0) PolarisBulk/0.2"},
        )
        if r.status_code != 200:
            return ""
        return extract_article_text(r.text)
    except Exception as e:
        logger.warning(f"fetch 실패 {url}: {e}")
        return ""

# ====== KRX — FinanceDataReader ======

def collect_krx_fdr(
    corps: list[dict], out_root: Path, manifest: "Manifest",
    from_year: int, to_year: int,
) -> None:
    """FinanceDataReader로 5사 일별 OHLCV 수집 (네이버/KRX 통합 소스)."""
    if fdr is None:
        logger.warning("FinanceDataReader 미설치 — pip install finance-datareader")
        manifest.skip(stage="krx", reason="fdr_not_installed")
        return
    today = date.today()
    end_yr = min(to_year, today.year)
    end = f"{end_yr}-12-31"
    if datetime.strptime(end, "%Y-%m-%d").date() > today:
        end = today.isoformat()
    logger.info(f"KRX(fdr) 범위: {from_year}-01-01 ~ {end}, 5사 {len(corps)} (증분: 과거연도 파일 있으면 올해만 재수집)")
    for c in corps:
        stock_code = c.get("stock_code", "")
        cc = c.get("corp_code", "")
        nm = c.get("corp_name", "?")
        if not stock_code or not cc:
            manifest.skip(stage="krx", corp=cc, reason="no_stock_code")
            continue
        # 증분: 과거연도 (today.year 이전) JSON 모두 있으면 올해치만 수집
        out = out_root / cc / "krx"
        out.mkdir(parents=True, exist_ok=True)
        past_years = list(range(from_year, today.year))
        past_all_present = all((out / f"daily_ohlcv_{y}.json").exists() for y in past_years)
        effective_start = (f"{today.year}-01-01" if past_all_present and past_years
                           else f"{from_year}-01-01")
        if effective_start != f"{from_year}-01-01":
            logger.info(f"  {nm}: 과거 {len(past_years)}년 cached → 올해 ({effective_start}~{end}) 만 갱신")
        try:
            df = fdr.DataReader(stock_code, effective_start, end)
        except Exception as e:
            logger.warning(f"KRX {nm} ({stock_code}) 실패: {e}")
            manifest.add(cc, "krx_ohlcv", "error")
            manifest.error(stage="krx", corp=cc, error=str(e))
            time.sleep(SLEEP_KRX_CORP)
            continue
        if df is None or df.empty:
            manifest.add(cc, "krx_ohlcv", "no_data")
            time.sleep(SLEEP_KRX_CORP)
            continue
        df = df.reset_index()
        date_col = df.columns[0]
        def _safe_float(v) -> float:
            """NaN / None / 빈값 → 0.0. 거래정지 일자 보호."""
            if v is None:
                return 0.0
            try:
                f = float(v)
            except (TypeError, ValueError):
                return 0.0
            return 0.0 if math.isnan(f) else f

        def _safe_int(v) -> int:
            if v is None:
                return 0
            try:
                f = float(v)
            except (TypeError, ValueError):
                return 0
            return 0 if math.isnan(f) else int(f)

        rows: list[dict] = []
        for _, r in df.iterrows():
            d = r[date_col]
            d_str = d.strftime("%Y%m%d") if hasattr(d, "strftime") else str(d)[:10].replace("-", "")
            rows.append({
                "basDd": d_str,
                "open": _safe_float(r.get("Open")),
                "high": _safe_float(r.get("High")),
                "low": _safe_float(r.get("Low")),
                "close": _safe_float(r.get("Close")),
                "volume": _safe_int(r.get("Volume")),
                "change_pct": _safe_float(r.get("Change")) * 100,
            })
        # 연도별 분할 저장
        by_year: dict[str, list[dict]] = defaultdict(list)
        for r in rows:
            by_year[r["basDd"][:4]].append(r)
        for y, rs in by_year.items():
            write_json(out / f"daily_ohlcv_{y}.json", {
                "year": int(y), "stock_code": stock_code,
                "corp_code": cc, "corp_name": nm, "rows": rs,
            })
        manifest.add(cc, "krx_ohlcv", "ok")
        logger.info(f"  {nm} ({stock_code}): {len(rows)} 일")
        time.sleep(SLEEP_KRX_CORP)

# ====== 카탈로그 ======

@dataclass
class CatalogEntry:
    category: str
    api_id: str
    endpoint: str
    name: str
    url: str
    required: list[str]
    optional: list[str]
    callable_: bool = True
    note: str = ""

def load_catalog(xlsx_path: Path) -> list[CatalogEntry]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["1.엔드포인트_메타"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    entries: list[CatalogEntry] = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        cat = str(row[0]).strip()
        api_id = str(row[1] or "").strip()
        ep = str(row[2] or "").strip()
        kname = str(row[3] or "").strip()
        url = str(row[4] or "").strip()
        req = str(row[5] or "")
        opt = str(row[6] or "")
        req_list = [x.strip() for x in req.split(",")
                    if x.strip() and x.strip() != "-" and x.strip() != "crtfc_key"]
        opt_list = [x.strip() for x in opt.split(",") if x.strip() and x.strip() != "-"]
        callable_ = url.startswith(("http://", "https://"))
        if not ep or ep.startswith("(") or "외" in ep:
            callable_ = False
        entries.append(CatalogEntry(
            category=cat, api_id=api_id, endpoint=ep, name=kname,
            url=url, required=req_list, optional=opt_list,
            callable_=callable_,
            note="" if callable_ else "non-callable meta",
        ))
    return entries

# ====== 회사 매핑 ======

def _build_corp_code_cache(dart: DartHTTPClient) -> dict[str, str]:
    """corpCode.xml ZIP → {stock_code: corp_code, stock_code+'_name': name}."""
    logger.info("DART corpCode.xml 1회 다운로드 (캐시 생성)")
    resp = dart.request("corpCode", {}, fmt="zip", url_ext="xml")
    if not isinstance(resp.data, (bytes, bytearray)):
        return {}
    with zipfile.ZipFile(io.BytesIO(resp.data)) as z:
        xml_bytes = z.read(z.namelist()[0])
    root = ET.fromstring(xml_bytes)
    cache: dict[str, str] = {}
    for li in root.iter("list"):
        cc = (li.findtext("corp_code") or "").strip()
        sc = (li.findtext("stock_code") or "").strip()
        nm = (li.findtext("corp_name") or "").strip()
        if sc:
            cache[sc] = cc
            cache[f"{sc}_name"] = nm
    return cache

def normalize_corps(
    corps_arg: str, corp_codes_arg: str,
    cache_path: Path, dart: DartHTTPClient,
) -> list[dict]:
    if not corps_arg and not corp_codes_arg:
        return [
            {"corp_code": cc, "stock_code": v["stock"], "corp_name": v["name"], "isin": v["isin"]}
            for cc, v in DEFAULT_CORPS.items()
        ]
    cache: dict = {}
    if cache_path.exists():
        try:
            cache = json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            cache = {}
    result: list[dict] = []
    seen: set[str] = set()

    def add(cc: str, sc: str, nm: str, isin: str = ""):
        if cc and cc not in seen:
            seen.add(cc)
            result.append({"corp_code": cc, "stock_code": sc, "corp_name": nm, "isin": isin})

    for cc in [x.strip() for x in (corp_codes_arg or "").split(",") if x.strip()]:
        if cc in DEFAULT_CORPS:
            v = DEFAULT_CORPS[cc]
            add(cc, v["stock"], v["name"], v["isin"])
        else:
            resp = dart.request("company", {"corp_code": cc})
            if resp.is_ok and isinstance(resp.data, dict):
                d = resp.data
                add(cc, d.get("stock_code", ""), d.get("corp_name", "?"))
            else:
                add(cc, "", "?")
    stocks = [x.strip() for x in (corps_arg or "").split(",") if x.strip()]
    stock_to_cc = {v["stock"]: cc for cc, v in DEFAULT_CORPS.items()}
    pending = [s for s in stocks if s not in stock_to_cc and s not in cache]
    if pending:
        cache.update(_build_corp_code_cache(dart))
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    for s in stocks:
        if s in stock_to_cc:
            cc = stock_to_cc[s]
            v = DEFAULT_CORPS[cc]
            add(cc, s, v["name"], v["isin"])
        elif s in cache:
            cc = cache[s]
            nm = cache.get(f"{s}_name", "?")
            add(cc, s, nm)
        else:
            logger.warning(f"stock {s} → corp_code 매핑 실패")
    return result

# ====== Manifest ======

def make_run_id(ts: str | None = None, seq: int = 1) -> str:
    """run_id 발급 — Pipeline_최종설계/05 §1.1 형식 `^[0-9]{8}_[0-9]{4}_[0-9]+$`.

    예: `20260524_1629_01`. ts None 이면 현재 시각으로 발급.
    """
    if ts is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
    elif "T" in ts:  # isoformat → yyyymmdd_HHMM
        ts = ts.replace("-", "").replace(":", "")[:13].replace("T", "_")
    return f"{ts}_{seq:02d}"


@dataclass
class Manifest:
    started: str
    finished: str = ""
    run_id: str = ""                # Pipeline_최종설계/05 §1.1 run_id (자동 발급)
    hash_algo: str = "sha1[:8]"     # 파일명 `__<hash>.json` 의 해시 알고리즘 (params_hash 참조)
    args: dict = field(default_factory=dict)
    corps: list[dict] = field(default_factory=list)
    by_corp: dict = field(default_factory=dict)
    by_endpoint: dict = field(default_factory=dict)
    errors: list[dict] = field(default_factory=list)
    skipped: list[dict] = field(default_factory=list)
    summary: dict = field(default_factory=dict)
    notes: dict = field(default_factory=dict)

    def add(self, corp_code: str, endpoint: str, status: str) -> None:
        self.by_corp.setdefault(corp_code, {}).setdefault(endpoint, {}).setdefault(status, 0)
        self.by_corp[corp_code][endpoint][status] += 1
        self.by_endpoint.setdefault(endpoint, {}).setdefault(status, 0)
        self.by_endpoint[endpoint][status] += 1

    def error(self, **kw) -> None:
        self.errors.append(kw)

    def skip(self, **kw) -> None:
        self.skipped.append(kw)

# ====== DART 수집 ======

def _save_dart_response(out_dir: Path, endpoint: str, params: dict, resp: DartResponse) -> Path:
    ph = params_hash(params)
    fn = f"{endpoint}__{ph}"
    out_dir.mkdir(parents=True, exist_ok=True)
    safe_params = {k: v for k, v in (resp.params or {}).items() if k != "crtfc_key"}
    if isinstance(resp.data, (bytes, bytearray)):
        (out_dir / f"{fn}.bin").write_bytes(resp.data)
        meta = {
            "endpoint": endpoint, "status": resp.status,
            "raw_status_code": resp.raw_status_code, "message": resp.message,
            "params": safe_params, "binary": True,
        }
        path = out_dir / f"{fn}.json"
        path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    else:
        payload = {
            "endpoint": endpoint, "status": resp.status,
            "raw_status_code": resp.raw_status_code, "message": resp.message,
            "params": safe_params, "data": resp.data,
        }
        path = out_dir / f"{fn}.json"
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return path

def _call_dart_safe(
    dart: DartHTTPClient, endpoint: str, params: dict,
    fmt: str = "json", url_ext: str | None = None,
) -> DartResponse | None:
    try:
        return dart.request(endpoint, params, fmt=fmt, url_ext=url_ext)
    except DartKeyError:
        raise
    except Exception as e:
        logger.warning(f"{endpoint} {params} 실패: {e}")
        return None

def collect_dart_endpoint(
    dart: DartHTTPClient, corps: list[dict], entry: CatalogEntry,
    out_root: Path, manifest: Manifest,
    from_year: int, to_year: int,
    rcept_no_pool: dict[str, list[str]],
) -> None:
    ep = entry.endpoint
    et = _ep_type(ep, entry.category)

    if et == "skip":
        manifest.skip(endpoint=ep, reason="unsupported_type", category=entry.category)
        return
    if et == "ftc":
        return

    if et == "no_arg_zip":
        out = out_root / "_common"
        out.mkdir(parents=True, exist_ok=True)
        try:
            resp = dart.request("corpCode", {}, fmt="zip", url_ext="xml")
            if isinstance(resp.data, (bytes, bytearray)):
                (out / "corpCode.zip").write_bytes(resp.data)
            manifest.add("_common", ep, resp.status)
        except DartKeyError:
            raise
        except Exception as e:
            manifest.error(endpoint=ep, error=str(e))
        return

    if et == "sj_div":
        for sj in ["BS1", "BS2", "BS3", "BS4", "IS1", "IS2", "IS3",
                   "CIS1", "CIS2", "DCF", "ICF", "SCE"]:
            params = {"sj_div": sj}
            resp = _call_dart_safe(dart, "xbrlTaxonomy", params)
            if resp is None:
                manifest.error(endpoint=ep, params=params, error="exception")
                continue
            _save_dart_response(out_root / "_common" / "xbrlTaxonomy", ep, params, resp)
            manifest.add("_common", ep, resp.status)
        return

    if et == "multi_corp_year_reprt":
        cc_csv = ",".join(c["corp_code"] for c in corps)
        out = out_root / "_common" / "fnlttMultiAcnt"
        for y in range(from_year, to_year + 1):
            for r in REPRT_CODES:
                params = {"corp_code": cc_csv, "bsns_year": str(y), "reprt_code": r}
                resp = _call_dart_safe(dart, "fnlttMultiAcnt", params)
                if resp is None:
                    manifest.error(endpoint=ep, params=params, error="exception")
                    continue
                _save_dart_response(out, ep, params, resp)
                manifest.add("_common", ep, resp.status)
        return

    if et == "multi_corp_year_reprt_idx":
        cc_csv = ",".join(c["corp_code"] for c in corps)
        out = out_root / "_common" / "fnlttCmpnyIndx"
        for y in range(from_year, to_year + 1):
            for r in REPRT_CODES:
                for idx in IDX_CL_CODES:
                    params = {"corp_code": cc_csv, "bsns_year": str(y),
                              "reprt_code": r, "idx_cl_code": idx}
                    resp = _call_dart_safe(dart, "fnlttCmpnyIndx", params)
                    if resp is None:
                        manifest.error(endpoint=ep, params=params, error="exception")
                        continue
                    _save_dart_response(out, ep, params, resp)
                    manifest.add("_common", ep, resp.status)
        return

    # 회사별
    for c in corps:
        cc = c["corp_code"]
        if not cc:
            continue
        out = out_root / cc / "dart"

        if et == "corp_code_only":
            params = {"corp_code": cc}
            resp = _call_dart_safe(dart, ep, params)
            if resp is None:
                manifest.error(corp=cc, endpoint=ep, error="exception")
                continue
            _save_dart_response(out, ep, params, resp)
            manifest.add(cc, ep, resp.status)

        elif et == "list":
            for y in range(from_year, to_year + 1):
                bgn = f"{y}0101"; end = f"{y}1231"
                page = 1
                while True:
                    params = {
                        "corp_code": cc, "bgn_de": bgn, "end_de": end,
                        "page_no": str(page), "page_count": "100",
                        "last_reprt_at": "Y",
                    }
                    resp = _call_dart_safe(dart, ep, params)
                    if resp is None:
                        manifest.error(corp=cc, endpoint=ep, year=y, page=page,
                                       error="exception")
                        break
                    _save_dart_response(out, ep, params, resp)
                    manifest.add(cc, ep, resp.status)
                    if not (resp.is_ok and isinstance(resp.data, dict)):
                        break
                    items = resp.data.get("list", []) or []
                    for item in items:
                        rno = item.get("rcept_no", "")
                        rn = item.get("report_nm", "")
                        if rno and _is_key_report(rn):
                            rcept_no_pool.setdefault(cc, []).append(rno)
                    total_page = int(resp.data.get("total_page", 1) or 1)
                    if page >= total_page or not items:
                        break
                    page += 1

        elif et == "corp_period":
            for y in range(from_year, to_year + 1):
                bgn = f"{y}0101"; end = f"{y}1231"
                params = {"corp_code": cc, "bgn_de": bgn, "end_de": end}
                resp = _call_dart_safe(dart, ep, params)
                if resp is None:
                    manifest.error(corp=cc, endpoint=ep, year=y, error="exception")
                    continue
                _save_dart_response(out, ep, params, resp)
                manifest.add(cc, ep, resp.status)

        elif et == "corp_year_reprt":
            for y in range(from_year, to_year + 1):
                for r in REPRT_CODES:
                    params = {"corp_code": cc, "bsns_year": str(y), "reprt_code": r}
                    resp = _call_dart_safe(dart, ep, params)
                    if resp is None:
                        manifest.error(corp=cc, endpoint=ep, year=y, reprt=r,
                                       error="exception")
                        continue
                    _save_dart_response(out, ep, params, resp)
                    manifest.add(cc, ep, resp.status)

        elif et == "corp_year_reprt_fs":
            for y in range(from_year, to_year + 1):
                for r in REPRT_CODES:
                    for fs in ["CFS", "OFS"]:
                        params = {"corp_code": cc, "bsns_year": str(y),
                                  "reprt_code": r, "fs_div": fs}
                        resp = _call_dart_safe(dart, ep, params)
                        if resp is None:
                            manifest.error(corp=cc, endpoint=ep, year=y, reprt=r, fs=fs,
                                           error="exception")
                            continue
                        _save_dart_response(out, ep, params, resp)
                        manifest.add(cc, ep, resp.status)

        elif et == "corp_year_reprt_idx":
            for y in range(from_year, to_year + 1):
                for r in REPRT_CODES:
                    for idx in IDX_CL_CODES:
                        params = {"corp_code": cc, "bsns_year": str(y),
                                  "reprt_code": r, "idx_cl_code": idx}
                        resp = _call_dart_safe(dart, ep, params)
                        if resp is None:
                            manifest.error(corp=cc, endpoint=ep, year=y, reprt=r, idx=idx,
                                           error="exception")
                            continue
                        _save_dart_response(out, ep, params, resp)
                        manifest.add(cc, ep, resp.status)

        elif et == "rcept_no":
            for rno in (rcept_no_pool.get(cc) or [])[:20]:
                params = {"rcept_no": rno}
                resp = _call_dart_safe(dart, "document", params, fmt="zip", url_ext="xml")
                if resp is None:
                    manifest.error(corp=cc, endpoint=ep, rno=rno, error="exception")
                    continue
                _save_dart_response(out, ep, params, resp)
                manifest.add(cc, ep, resp.status)

        elif et == "rcept_no_reprt":
            for rno in (rcept_no_pool.get(cc) or [])[:10]:
                for r in ["11011"]:
                    params = {"rcept_no": rno, "reprt_code": r}
                    resp = _call_dart_safe(dart, "fnlttXbrl", params, fmt="zip", url_ext="xml")
                    if resp is None:
                        manifest.error(corp=cc, endpoint=ep, rno=rno, reprt=r,
                                       error="exception")
                        continue
                    _save_dart_response(out, ep, params, resp)
                    manifest.add(cc, ep, resp.status)

def collect_dart(
    dart: DartHTTPClient, corps: list[dict], catalog: list[CatalogEntry],
    out_root: Path, manifest: Manifest, from_year: int, to_year: int,
) -> dict[str, list[str]]:
    rcept_no_pool: dict[str, list[str]] = {}
    dart_eps = [e for e in catalog if e.callable_ and e.category.startswith("DS")]
    dart_eps.sort(key=lambda e: 0 if e.endpoint == "list"
                  else 1 if e.endpoint == "company"
                  else 2 if e.endpoint == "corpCode"
                  else 10)
    for entry in dart_eps:
        logger.info(f"DART {entry.category} {entry.endpoint} ({entry.name})")
        try:
            collect_dart_endpoint(dart, corps, entry, out_root, manifest,
                                  from_year, to_year, rcept_no_pool)
        except DartKeyError:
            raise
        except Exception as e:
            logger.error(f"endpoint {entry.endpoint} 실패: {e}")
            manifest.error(endpoint=entry.endpoint, error=str(e))
    return rcept_no_pool

def rebuild_pool_from_list_files(out_root: Path, corps: list[dict]) -> dict[str, list[str]]:
    """기존 list 응답 파일에서 rcept_no_pool 재구성."""
    pool: dict[str, list[str]] = {}
    for c in corps:
        cc = c["corp_code"]
        if not cc:
            continue
        dart_dir = out_root / cc / "dart"
        if not dart_dir.exists():
            continue
        rnos: list[str] = []
        for f in dart_dir.glob("list__*.json"):
            try:
                obj = json.loads(f.read_text(encoding="utf-8"))
                items = (obj.get("data", {}) or {}).get("list") or []
                for item in items:
                    rno = item.get("rcept_no", "")
                    rn = item.get("report_nm", "")
                    if rno and _is_key_report(rn):
                        rnos.append(rno)
            except Exception:
                pass
        if rnos:
            pool[cc] = list(dict.fromkeys(rnos))
    return pool

# ====== 사업보고서 원문 ======

def collect_documents(
    dart: DartHTTPClient, corps: list[dict], rcept_no_pool: dict[str, list[str]],
    out_root: Path, manifest: Manifest, max_per_corp: int = 300,
) -> None:
    """사업/반기/분기보고서 + 주요 공시 원문 PDF/HTML 다운로드."""
    pdf_http = httpx.Client(
        timeout=60, follow_redirects=True,
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0) PolarisBulk/0.2"},
    )
    try:
        for ci, c in enumerate(corps):
            cc = c["corp_code"]
            nm = c.get("corp_name", "?")
            if not cc:
                continue
            rnos = list(dict.fromkeys(rcept_no_pool.get(cc, [])))[:max_per_corp]
            if not rnos:
                manifest.skip(corp=cc, reason="no_rcept_no_in_pool")
                continue
            logger.info(f"documents {nm} ({cc}) — rno {len(rnos)}건")
            for ri, rno in enumerate(rnos):
                out = out_root / cc / "documents" / rno
                existing_html = (out / "body.html").exists()
                existing_pdf = (out / "original.pdf").exists()
                if existing_html and existing_pdf:
                    manifest.add(cc, "documents", "cached")
                    continue

                if existing_html and not existing_pdf:
                    # PDF만 재시도 (ZIP 재호출 X). main.do에 pdf.do 링크 없으면 즉시 skip.
                    success = False
                    no_pdf_available = False
                    for backoff in PDF_RETRY_BACKOFF:
                        time.sleep(backoff)
                        try:
                            pdf_url, referer = find_pdf_url(pdf_http, rno)
                            if pdf_url is None:
                                no_pdf_available = True
                                break
                            if pdf_url:
                                time.sleep(SLEEP_PDF_BETWEEN)
                                if download_pdf(pdf_http, rno, pdf_url, out / "original.pdf", referer):
                                    success = True
                                    meta = {}
                                    if (out / "meta.json").exists():
                                        try:
                                            meta = json.loads((out / "meta.json").read_text(encoding="utf-8"))
                                        except Exception:
                                            meta = {}
                                    meta["pdf"] = str(out / "original.pdf")
                                    meta["pdf_url"] = pdf_url
                                    write_json(out / "meta.json", meta)
                                    break
                        except Exception:
                            pass
                    if success:
                        manifest.add(cc, "documents", "pdf_added")
                    elif no_pdf_available:
                        manifest.add(cc, "documents", "no_pdf_html_only")
                    else:
                        manifest.add(cc, "documents", "pdf_fail")
                    time.sleep(SLEEP_RNO_GAP)
                    continue

                # 처음부터 (ZIP + body + PDF)
                try:
                    resp = dart.request("document", {"rcept_no": rno}, fmt="zip", url_ext="xml")
                except DartKeyError:
                    raise
                except Exception as e:
                    logger.warning(f"document ZIP 실패 {rno}: {e}")
                    manifest.error(corp=cc, rno=rno, error=str(e))
                    continue
                if not isinstance(resp.data, (bytes, bytearray)):
                    manifest.add(cc, "documents", resp.status)
                    write_json(out / "meta.json", {
                        "rcept_no": rno, "corp_code": cc,
                        "status": resp.status, "message": resp.message,
                    })
                    continue
                try:
                    dl = fetch_business_report(bytes(resp.data), rno, cc, out, pdf_http)
                except Exception as e:
                    logger.warning(f"fetch_business_report 실패 {rno}: {e}")
                    manifest.error(corp=cc, rno=rno, error=str(e))
                    continue
                write_json(out / "meta.json", {
                    "rcept_no": rno, "corp_code": cc,
                    "body_html": str(dl.body_html_path) if dl.body_html_path else None,
                    "pdf": str(dl.pdf_path) if dl.pdf_path else None,
                    "pdf_url": dl.pdf_url,
                    "error": dl.error,
                })
                if dl.pdf_path and dl.body_html_path:
                    manifest.add(cc, "documents", "ok")
                elif dl.body_html_path:
                    manifest.add(cc, "documents", "html_only")
                else:
                    manifest.add(cc, "documents", "fail")
                time.sleep(SLEEP_RNO_GAP)
            # 회사 사이 충분히 쉬기 (DART 뷰어 IP 차단 방지)
            if ci < len(corps) - 1:
                next_nm = corps[ci + 1].get("corp_name", "?")
                logger.info(f"  ★ 회사 사이 {SLEEP_CORP_GAP:.0f}s 대기 ({ci + 1}/{len(corps)} 완료, 다음 → {next_nm})")
                time.sleep(SLEEP_CORP_GAP)
    finally:
        pdf_http.close()

# ====== 뉴스 ======

def collect_news_wrap(
    corps: list[dict], out_root: Path, manifest: Manifest,
    max_per_feed: int, since_date: date | None,
) -> None:
    """knews-rss RSS 전체 기사 수집 (since_date 이후만).

    회사 매칭 없이 모든 기사를 _common/news/{news_id}.json에만 저장.
    향후 분석 시 corp_code 매핑이 필요하면 별도 단계로 처리.
    """
    common_out = out_root / "_common" / "news"
    common_out.mkdir(parents=True, exist_ok=True)
    since_str = since_date.isoformat() if since_date else "(no filter)"
    logger.info(f"뉴스 수집 — since={since_str}")

    http = httpx.Client(
        timeout=30, follow_redirects=True,
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0) PolarisBulk/0.2"},
    )
    stats = {
        "feeds_total": 0, "feeds_selected": 0,
        "entries_total": 0, "entries_filtered_old": 0,
        "entries_fetched": 0, "entries_failed": 0,
        "entries_saved": 0, "entries_cached": 0,
    }
    try:
        feed_specs = fetch_feed_specs(http)
        stats["feeds_total"] = len(feed_specs)
        selected = select_feeds(feed_specs)
        stats["feeds_selected"] = len(selected)

        for fi, feed in enumerate(selected, 1):
            url = feed.get("url")
            if not url:
                continue
            parsed = fetch_feed_parsed(http, url)
            entries = parsed.entries[:max_per_feed]
            stats["entries_total"] += len(entries)
            if fi % 5 == 0:
                logger.info(f"  feed {fi}/{len(selected)} — 누적 saved={stats['entries_saved']}")

            for entry in entries:
                e_url = entry.get("link")
                if not e_url:
                    continue
                if since_date:
                    pub_t = entry.get("published_parsed") or entry.get("updated_parsed")
                    if pub_t:
                        try:
                            pub_dt = datetime.fromtimestamp(time.mktime(pub_t)).date()
                            if pub_dt < since_date:
                                stats["entries_filtered_old"] += 1
                                continue
                        except Exception:
                            pass
                nid = news_id_from_url(e_url)
                target = common_out / f"{nid}.json"
                if target.exists():
                    stats["entries_cached"] += 1
                    manifest.add("_common", "news_all", "cached")
                    continue
                body = fetch_article_body(http, e_url)
                if not body:
                    stats["entries_failed"] += 1
                    manifest.add("_common", "news_all", "fetch_fail")
                    time.sleep(SLEEP_NEWS_ENTRY)
                    continue
                stats["entries_fetched"] += 1
                title = entry.get("title", "")
                published = entry.get("published", "") or entry.get("updated", "")
                item = NewsItem(
                    news_id=nid, url=e_url, title=title, published=published,
                    source_feed=url, source_publisher=feed.get("publisher", ""),
                    text=body,
                )
                # 파일 (백업) + RDB news_raw (SSOT) 동시 저장
                target.write_text(
                    json.dumps(item.__dict__, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                _news_raw_insert(item, feed)
                stats["entries_saved"] += 1
                manifest.add("_common", "news_all", "ok")
                time.sleep(SLEEP_NEWS_ENTRY)
        manifest.notes["news_stats"] = stats
        manifest.notes["news_since"] = since_str
        logger.info(f"뉴스 수집 완료: {stats}")
    except Exception as e:
        logger.error(f"뉴스 수집 실패: {e}")
        manifest.error(stage="news", error=str(e))
    finally:
        http.close()

# ====== FTC (공정위) ======
# G4: catalog/1.엔드포인트_메타에서 cat=='FTC' 동적 로드 + 캐시 적용
# base URL은 카탈로그에 들어있음 (예: http://apis.data.go.kr/1130000/{stem}/{stem}Api)
# 호출 흐름: (1) appnGroupSttusList로 그룹 목록 → (2) 5사 관련 그룹 추출 → (3) 나머지 endpoint별 호출

FTC_BASE_URL = "http://apis.data.go.kr/1130000"

# Backward compat (다른 모듈/스크립트가 이름 import할 수 있음)
FTC_GROUP_URL = f"{FTC_BASE_URL}/appnGroupSttusList/appnGroupSttusListApi"
FTC_AFLT_URL = f"{FTC_BASE_URL}/appnGroupAffiList/appnGroupAffiListApi"
FTC_INDTY_URL = f"{FTC_BASE_URL}/typeOfBusinessCompSttusList/typeOfBusinessCompSttusListApi"

# 인증키 대소문자 — endpoint별 상이 (spec/8 비고)
FTC_KEY_CASE = {
    'appnGroupSttusList': 'serviceKey',
    'appnGroupAffiList': 'ServiceKey',
    'typeOfBusinessCompSttusList': 'ServiceKey',
    'affiliationCompSttusList': 'ServiceKey',
    'executiveCompSttusList': 'ServiceKey',
    'stockholderCompSttusList': 'ServiceKey',
    'financeCompSttusList': 'ServiceKey',
    'sllInnerQotaList': 'serviceKey',
    'grupRotatInvstmntList': 'serviceKey',
    'innerQotaEqltrmCmprUnityList': 'serviceKey',
    'innerQotaEqltrmCmprAssetsList': 'serviceKey',
    'holdingProgCompStusList': 'serviceKey',
    'holdingGenFinCompSttusList': 'ServiceKey',
    'tyAssetsRentDelngDtlsList': 'serviceKey',
}

def _ftc_cache_path(out_root: Path, stem: str, params: dict) -> Path:
    """FTC 캐시 키 = stem + params hash (serviceKey 제외)."""
    safe = {k: v for k, v in params.items() if k not in ('serviceKey', 'ServiceKey')}
    ph = params_hash(safe)
    cache_dir = out_root / "_common" / "ftc_cache" / stem
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / f"{stem}__{ph}.xml"

def _ftc_fetch_xml(http: httpx.Client, url: str, params: dict,
                   cache_path: Path | None = None, fresh: bool = False) -> tuple[int, str, bool]:
    """Returns (http_status, body_text, cached). DART과 동일 캐시 패턴."""
    if cache_path and cache_path.exists() and not fresh:
        return 200, cache_path.read_text(encoding='utf-8'), True
    try:
        r = http.get(url, params=params)
        time.sleep(SLEEP_FTC_REQUEST)
        if cache_path and r.status_code == 200:
            cache_path.write_text(r.text, encoding='utf-8')
        return r.status_code, r.text, False
    except Exception as e:
        return 0, f"exception:{e}", False

def _ftc_parse_groups(xml_text: str) -> tuple[str, int, int, list[dict]]:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return ("parse_error", 0, 0, [])
    rc = root.findtext("resultCode", "")
    tc = int(root.findtext("totalCount", "0") or 0)
    nr = int(root.findtext("numOfRows", "0") or 0)
    groups = [
        {c.tag: (c.text or "") for c in it}
        for it in root.findall("appnGroupSttus")
    ]
    return rc, tc, nr, groups

def _ftc_count_items(xml_text: str, item_tag: str) -> tuple[str, int, int, int]:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return ("parse_error", 0, 0, 0)
    rc = root.findtext("resultCode", "")
    tc = int(root.findtext("totalCount", "0") or 0)
    nr = int(root.findtext("numOfRows", "0") or 0)
    n = len(root.findall(item_tag))
    return rc, tc, nr, n

# G4: FTC endpoint별 호출 패턴 정의 (필수 파라미터 → 함수)
# 그룹 단위 = unityGrupCode 필요, 기간 단위 = presentnYm/presentnYear 필요
FTC_ENDPOINT_PATTERN = {
    # 1. 그룹 목록 (시작점)
    'appnGroupSttusList': {'mode': 'year_only', 'param_year': 'presentnYear'},
    # 2. 소속회사 시리즈 (year + group)
    'appnGroupAffiList':           {'mode': 'year_group', 'param_year': 'presentnYear'},
    'typeOfBusinessCompSttusList': {'mode': 'year_group', 'param_year': 'presentnYear'},
    'affiliationCompSttusList':    {'mode': 'year_group', 'param_year': 'presentnYear'},
    'executiveCompSttusList':      {'mode': 'year_group', 'param_year': 'presentnYear'},
    'stockholderCompSttusList':    {'mode': 'year_group', 'param_year': 'presentnYear'},
    'financeCompSttusList':        {'mode': 'year_group', 'param_year': 'presentnYear'},
    # 3. 그룹 단위 (presentnYm + 그룹 또는 단독)
    'sllInnerQotaList':                  {'mode': 'ym_only', 'param_ym': 'presentnYm'},
    'grupRotatInvstmntList':             {'mode': 'ym_only', 'param_ym': 'presentnYm'},
    'innerQotaEqltrmCmprUnityList':      {'mode': 'group_only'},  # serviceKey, unityGrupCode
    'innerQotaEqltrmCmprAssetsList':     {'mode': 'ym_only', 'param_ym': 'presentnYm'},
    # 4. 지주회사 (jurirno 필요 — 5사 jurirno 필요)
    'holdingProgCompStusList':    {'mode': 'ym_jurirno', 'param_ym': 'presentnYm'},
    'holdingGenFinCompSttusList': {'mode': 'ym_only', 'param_ym': 'presentnYm'},
    # 5. 기간 단위
    'tyAssetsRentDelngDtlsList': {'mode': 'period_group',
                                   'param_begin': 'pdBeginPresentnYm', 'param_end': 'pdEndPresentnYm'},
}

def _ftc_endpoint_url(stem: str) -> str:
    """카탈로그 비정규 URL 처리 + 기본 패턴."""
    # 비정규 URL — spec/8 검증
    if stem == 'innerQotaEqltrmCmprUnityList' or stem == 'innerQotaEqltrmCmprAssetsList':
        return f"{FTC_BASE_URL}/innerQotaEqltrmCmprList/{stem}Api"
    if stem == 'holdingProgCompStusList':
        return f"{FTC_BASE_URL}/holdingProgCompSttusList/{stem}Api"  # path1=Sttus(t포함)
    # 표준 패턴
    return f"{FTC_BASE_URL}/{stem}/{stem}Api"

def _ftc_service_key(stem: str, service_key: str) -> dict:
    """endpoint별 대소문자 인증키 파라미터."""
    key_name = FTC_KEY_CASE.get(stem, 'serviceKey')
    return {key_name: service_key}

def _ftc_call_endpoint(http, stem: str, base_params: dict, out_root: Path,
                       manifest: Manifest, label_suffix: str = "") -> None:
    """단일 endpoint 호출 + 페이징 + 캐시 + manifest."""
    url = _ftc_endpoint_url(stem)
    out_dir = out_root / "_common" / "ftc" / stem
    out_dir.mkdir(parents=True, exist_ok=True)
    page = 1
    while True:
        params = dict(base_params)
        params['pageNo'] = str(page)
        params['numOfRows'] = '100'
        cache_p = _ftc_cache_path(out_root, stem, params)
        code, body_text, cached = _ftc_fetch_xml(http, url, params, cache_path=cache_p)
        if code != 200:
            manifest.add("_common", f"ftc_{stem}", f"http_{code}")
            break
        # 저장 (캐시와 별개로 사람이 보기 좋게 stem별 디렉토리)
        fn = f"{stem}{label_suffix}_p{page}.xml"
        (out_dir / fn).write_text(body_text, encoding='utf-8')
        # 응답 파싱 (resultCode + 페이지 정보)
        try:
            root = ET.fromstring(body_text)
            rc = root.findtext('resultCode', '')
            tc = int(root.findtext('totalCount', '0') or 0)
            nr = int(root.findtext('numOfRows', '0') or 0)
        except ET.ParseError:
            manifest.add("_common", f"ftc_{stem}", "parse_error")
            break
        if rc != "00":
            status = "no_data" if rc == "97" else f"rc_{rc}"
            manifest.add("_common", f"ftc_{stem}", status)
            break
        manifest.add("_common", f"ftc_{stem}", "cached" if cached else "ok")
        if page * (nr or 100) >= tc or tc == 0:
            break
        page += 1

def collect_ftc(out_root: Path, manifest: Manifest, corps: list[dict],
                from_year: int, to_year: int,
                catalog_entries: list[CatalogEntry] | None = None,
                ftc_endpoints_filter: list[str] | None = None) -> None:
    """공정위 OpenAPI — 카탈로그 동적 로드 + 14 endpoint 처리.

    Args:
        catalog_entries: load_catalog 결과. None이면 FTC_ENDPOINT_PATTERN 키 기준.
        ftc_endpoints_filter: 호출할 stem 목록 제한 (None=전부).
    """
    if not FTC_OPENAPI_KEY:
        logger.warning("FTC_OPENAPI_KEY가 없음 — FTC 스킵")
        manifest.skip(stage="ftc", reason="no_key")
        return
    service_key = unquote(FTC_OPENAPI_KEY)
    http = httpx.Client(timeout=30, follow_redirects=True,
                        headers={"User-Agent": "PolarisBulk/0.2"})

    # 카탈로그에서 FTC endpoint 목록 (callable만)
    if catalog_entries:
        cat_ftc = [e.endpoint for e in catalog_entries
                   if e.category == 'FTC' and e.callable_]
    else:
        cat_ftc = list(FTC_ENDPOINT_PATTERN.keys())
    if ftc_endpoints_filter:
        cat_ftc = [e for e in cat_ftc if e in ftc_endpoints_filter]
    # FTC_ENDPOINT_PATTERN에 정의 없는 endpoint는 skip + 로그
    callable_stems = [e for e in cat_ftc if e in FTC_ENDPOINT_PATTERN]
    unknown_stems = [e for e in cat_ftc if e not in FTC_ENDPOINT_PATTERN]
    if unknown_stems:
        for s in unknown_stems:
            manifest.skip(stage="ftc", endpoint=s, reason="pattern_undefined")
        logger.warning(f"FTC pattern 미정의 (skip): {unknown_stems}")

    try:
        # === 1단계: 그룹 목록 (appnGroupSttusList) ===
        groups_by_year: dict[int, list[dict]] = {}
        stem = 'appnGroupSttusList'
        url = _ftc_endpoint_url(stem)
        grp_out = out_root / "_common" / "ftc" / stem
        grp_out.mkdir(parents=True, exist_ok=True)
        logger.info(f"FTC 1단계: 그룹 목록 ({stem}, {from_year}~{to_year})")
        for y in range(from_year, to_year + 1):
            page = 1
            while True:
                params = {**_ftc_service_key(stem, service_key),
                          'presentnYear': str(y), 'pageNo': str(page), 'numOfRows': '100'}
                cache_p = _ftc_cache_path(out_root, stem, params)
                code, body_text, cached = _ftc_fetch_xml(http, url, params, cache_path=cache_p)
                if code != 200:
                    manifest.add("_common", f"ftc_{stem}", f"http_{code}")
                    break
                (grp_out / f"{stem}__{y}_p{page}.xml").write_text(body_text, encoding="utf-8")
                rc, tc, nr, items = _ftc_parse_groups(body_text)
                if rc != "00":
                    manifest.add("_common", f"ftc_{stem}", "no_data" if rc == "97" else f"rc_{rc}")
                    break
                manifest.add("_common", f"ftc_{stem}", "cached" if cached else "ok")
                groups_by_year.setdefault(y, []).extend(items)
                if page * (nr or 100) >= tc or not items:
                    break
                page += 1

        # === 2단계: 5사 관련 그룹 추출 ===
        corp_names = {c["corp_name"] for c in corps if c.get("corp_name")}
        kw = corp_names | {"삼성", "SK", "에스케이", "SK하이닉스"}
        target_groups: set[tuple[int, str, str]] = set()
        for y, gs in groups_by_year.items():
            for g in gs:
                gnm = g.get("unityGrupNm", "")
                rpr = g.get("repreCmpny", "")
                gcd = g.get("unityGrupCode", "")
                if gcd and any(k in gnm or k in rpr for k in kw):
                    target_groups.add((y, gcd, gnm))
        logger.info(f"FTC 5사 관련 그룹 {len(target_groups)}개")

        # === 3단계: endpoint별 호출 ===
        for stem in callable_stems:
            if stem == 'appnGroupSttusList':
                continue  # 1단계에서 처리
            pat = FTC_ENDPOINT_PATTERN[stem]
            mode = pat['mode']
            logger.info(f"FTC: {stem} (mode={mode})")
            if mode == 'year_group':
                # 5사 그룹 × 연도
                for y, gcd, gnm in target_groups:
                    base = {**_ftc_service_key(stem, service_key),
                            pat['param_year']: str(y), 'unityGrupCode': gcd}
                    _ftc_call_endpoint(http, stem, base, out_root, manifest,
                                       label_suffix=f"__{y}_{gcd}")
            elif mode == 'ym_only':
                # 연도×월 (YYYYMM) — 매년 12월 기준
                for y in range(from_year, to_year + 1):
                    base = {**_ftc_service_key(stem, service_key),
                            pat['param_ym']: f"{y}12"}
                    _ftc_call_endpoint(http, stem, base, out_root, manifest,
                                       label_suffix=f"__{y}12")
            elif mode == 'group_only':
                # unityGrupCode만
                for y, gcd, gnm in target_groups:
                    base = {**_ftc_service_key(stem, service_key),
                            'unityGrupCode': gcd}
                    _ftc_call_endpoint(http, stem, base, out_root, manifest,
                                       label_suffix=f"__{gcd}")
            elif mode == 'ym_jurirno':
                # presentnYm + jurirno (5사 jurirno 필요 — corps에 없으면 skip)
                for c in corps:
                    jur = c.get('jurirno') or ''
                    if not jur:
                        continue
                    for y in range(from_year, to_year + 1):
                        base = {**_ftc_service_key(stem, service_key),
                                pat['param_ym']: f"{y}12", 'jurirno': jur}
                        _ftc_call_endpoint(http, stem, base, out_root, manifest,
                                           label_suffix=f"__{y}12_{jur}")
            elif mode == 'period_group':
                # 기간(YYYYMMDD) + group
                for y, gcd, gnm in target_groups:
                    base = {**_ftc_service_key(stem, service_key),
                            pat['param_begin']: f"{y}0101",
                            pat['param_end']: f"{y}1231",
                            'unityGrupCode': gcd}
                    _ftc_call_endpoint(http, stem, base, out_root, manifest,
                                       label_suffix=f"__{y}_{gcd}")
    finally:
        http.close()

# ====== BOK ECOS (한국은행 경제통계) ======
# https://ecos.bok.or.kr/api
# 핵심 지표 시계열 — Pipeline 02 §2: "BOK 핵심 지표 ~30종 × 분기 ≈ 720"

BOK_BASE = "https://ecos.bok.or.kr/api"

# 핵심 지표 통계표 코드 (BOK 100대 통계지표 + Polaris 분석용)
# 2026-05-24: 4건 코드 정정 (731Y001→731Y004 환율, 901Y011/Y012→301Y013 수출입, 200Y001→200Y104 GDP)
BOK_INDICATORS = [
    # 통화·금리
    ("722Y001", "한국은행 기준금리", "M"),
    ("817Y002", "시장금리 (CD 91일)", "D"),
    # 물가
    ("901Y009", "소비자물가지수 (총지수)", "M"),
    ("901Y010", "근원물가지수 (식품·에너지제외)", "M"),
    ("404Y014", "생산자물가지수", "M"),
    # 경기·생산
    ("901Y033", "전산업 생산지수", "M"),
    ("901Y027", "광공업 생산지수", "M"),
    # 환율 (정정: 731Y001→731Y004 매매기준율)
    ("731Y004", "원/달러 환율 (매매기준율)", "M"),
    # 고용
    ("901Y084", "실업률 (계절조정)", "M"),
    # 무역·국제수지 (정정: 901Y011/Y012→301Y013 경상수지·관세청)
    ("301Y013", "경상수지·관세청 수출입 통합", "M"),
    ("902Y004", "수출입금액 (월별)", "M"),
    # GDP (정정: 200Y001→200Y104 국민계정)
    ("200Y104", "GDP (실질, 분기)", "Q"),
]

def collect_bok(out_root: Path, manifest: Manifest, from_year: int, to_year: int) -> None:
    """BOK ECOS — 핵심 지표 시계열 다운로드."""
    if not BOK_ECOS_KEY:
        logger.warning("BOK_ECOS_KEY 없음 — BOK 스킵")
        manifest.skip(stage="bok", reason="no_key")
        return
    out_dir = out_root / "_common" / "bok"
    out_dir.mkdir(parents=True, exist_ok=True)
    http = httpx.Client(timeout=30, follow_redirects=True,
                        headers={"User-Agent": "PolarisBulk/0.2"})
    try:
        for stat_code, name, freq in BOK_INDICATORS:
            # 주기별 시작/종료 형식
            if freq == "M":
                bgn, end = f"{from_year}01", f"{to_year}12"
            elif freq == "Q":
                bgn, end = f"{from_year}Q1", f"{to_year}Q4"
            elif freq == "D":
                bgn, end = f"{from_year}0101", f"{to_year}1231"
            else:
                bgn, end = f"{from_year}", f"{to_year}"
            # ECOS API URL
            url = (f"{BOK_BASE}/StatisticSearch/{BOK_ECOS_KEY}/json/kr/"
                   f"1/10000/{stat_code}/{freq}/{bgn}/{end}")
            cache_p = out_dir / f"{stat_code}_{freq}_{bgn}_{end}.json"
            if cache_p.exists():
                manifest.add("_common", f"bok_{stat_code}", "cached")
                continue
            try:
                r = http.get(url, timeout=30)
                time.sleep(SLEEP_FTC_REQUEST)
                if r.status_code != 200:
                    manifest.add("_common", f"bok_{stat_code}", f"http_{r.status_code}")
                    continue
                data = r.json()
                # ECOS는 RESULT.CODE 정상=INFO-000
                result = (data.get("StatisticSearch") or {}).get("list_total_count")
                if data.get("RESULT", {}).get("CODE") and data["RESULT"]["CODE"] != "INFO-000":
                    manifest.add("_common", f"bok_{stat_code}",
                                 f"err_{data['RESULT']['CODE']}")
                    continue
                cache_p.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                                  encoding="utf-8")
                manifest.add("_common", f"bok_{stat_code}", "ok")
                logger.info(f"  BOK {stat_code} ({name}) [{freq}]: {result} 행")
            except Exception as e:
                logger.warning(f"BOK {stat_code} 실패: {e}")
                manifest.add("_common", f"bok_{stat_code}", "error")
    finally:
        http.close()

# ====== KOSIS (통계청 국가통계포털) ======
# https://kosis.kr/openapi
# 핵심 거시통계 — 인구·고용·물가·산업

KOSIS_LIST_BASE = "https://kosis.kr/openapi/statisticsList.do"

# 핵심 거시 카테고리 (parentListId — KOSIS 주제별 목록)
# 2026-05-24: 통계자료(statisticsParameterData) endpoint는 통계표별 KOSIS 활용신청 필수.
# 신청 안 한 통계표는 err_21 → 메타 목록(statisticsList)만 받음. 데이터는 BOK ECOS 대체.
KOSIS_PARENT_CATEGORIES = [
    ("A", "인구·가구"),
    ("B", "노동"),
    ("C", "소득·소비·자산"),
    ("D", "보건·사회"),
    ("F", "교육·훈련"),
    ("J", "농림·수산"),
    ("K", "광공업·에너지"),
    ("L", "건설·주택·국토"),
    ("M", "교통·물류·정보통신"),
    ("N", "무역·외환·국제수지"),
    ("O", "기업경영"),
    ("P", "금융"),
    ("Q", "재정"),
    ("R", "물가"),
    ("S", "환경"),
]

def collect_kosis(out_root: Path, manifest: Manifest,
                  from_year: int, to_year: int) -> None:
    """KOSIS — 통계표 메타 목록 (statisticsList).
    통계자료(데이터)는 통계표별 활용신청 필수 → 별도 작업 필요.
    """
    if not KOSIS_KEY:
        logger.warning("KOSIS_KEY 없음 — KOSIS 스킵")
        manifest.skip(stage="kosis", reason="no_key")
        return
    out_dir = out_root / "_common" / "kosis"
    out_dir.mkdir(parents=True, exist_ok=True)
    http = httpx.Client(timeout=30, follow_redirects=True,
                        headers={"User-Agent": "PolarisBulk/0.2"})
    try:
        for parent_id, label in KOSIS_PARENT_CATEGORIES:
            params = {
                "method": "getList",
                "apiKey": KOSIS_KEY,
                "vwCd": "MT_ZTITLE",
                "parentListId": parent_id,
                "format": "json",
                "jsonVD": "Y",
            }
            cache_p = out_dir / f"list_{parent_id}.json"
            if cache_p.exists():
                manifest.add("_common", f"kosis_list_{parent_id}", "cached")
                continue
            try:
                r = http.get(KOSIS_LIST_BASE, params=params, timeout=30)
                time.sleep(SLEEP_FTC_REQUEST)
                if r.status_code != 200:
                    manifest.add("_common", f"kosis_list_{parent_id}", f"http_{r.status_code}")
                    continue
                try:
                    data = json.loads(r.text)
                except json.JSONDecodeError:
                    manifest.add("_common", f"kosis_list_{parent_id}", "parse_error")
                    continue
                if isinstance(data, dict) and data.get("err"):
                    manifest.add("_common", f"kosis_list_{parent_id}", f"err_{data.get('err')}")
                    continue
                cache_p.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                                  encoding="utf-8")
                rows = len(data) if isinstance(data, list) else 1
                manifest.add("_common", f"kosis_list_{parent_id}", "ok")
                logger.info(f"  KOSIS list [{parent_id} {label}]: {rows} 목록")
            except Exception as e:
                logger.warning(f"KOSIS list {parent_id} 실패: {e}")
                manifest.add("_common", f"kosis_list_{parent_id}", "error")
    finally:
        http.close()

# ====== Manifest 마무리 ======

def finalize_manifest(manifest: Manifest, out_root: Path) -> None:
    manifest.finished = datetime.now().isoformat(timespec="seconds")
    total_ok = sum(
        sum(v for k, v in eps.items() if k == "ok")
        for ep_map in manifest.by_corp.values()
        for eps in ep_map.values()
    )
    total = sum(
        sum(eps.values())
        for ep_map in manifest.by_corp.values()
        for eps in ep_map.values()
    )
    manifest.summary = {
        "total_calls": total,
        "ok": total_ok,
        "ok_ratio": round(total_ok / total, 3) if total else 0,
        "errors": len(manifest.errors),
        "skipped": len(manifest.skipped),
    }
    manifest.notes.setdefault("hankyung_maekyung",
                              "skipped (no crawler in zzzf or knews-rss)")
    write_json(out_root / "_manifest.json", asdict(manifest))

# ====== CLI ======

@app.command()
def main(
    corps: str = typer.Option("", help="stock_code 콤마. 비우면 기본 5사."),
    corp_codes: str = typer.Option("", help="corp_code 콤마. corps와 동시 사용 가능."),
    from_year: int = typer.Option(2025, help="시작 사업연도"),
    to_year: int = typer.Option(0, help="종료 사업연도 (0이면 올해)"),
    news_max_per_feed: int = typer.Option(200, help="RSS 피드당 최대 entry"),
    news_since: str = typer.Option("2026-01-01", help="이 날짜 이후 published만 (YYYY-MM-DD)"),
    docs_max_per_corp: int = typer.Option(300, help="회사당 공시 원문 다운로드 상한"),
    only: str = typer.Option("", help="dart,documents,krx,news,ftc 중 콤마"),
    skip: str = typer.Option("", help="dart,documents,krx,news,ftc 중 제외"),
    out_dir: str = typer.Option(str(DEFAULT_OUT), help="산출물 디렉토리"),
    catalog: str = typer.Option(str(DEFAULT_CATALOG), help="카탈로그 xlsx 경로"),
    profile: str = typer.Option("normal", help="텀 프로필: slow(30사+ 보수), normal(현재값), fast(개발)"),
    no_rebuild_pool: bool = typer.Option(False, help="rcept_no_pool 강제 유지 — list 응답 기반 재구성 안 함 (PDF 재시도 등 특정 rno만 처리 시)"),
):
    apply_rate_profile(profile)
    """카탈로그 기반 5사(+추가) 일괄 수집기 — 자족 단일 파일."""
    today = date.today()
    if to_year == 0:
        to_year = today.year

    out_root = Path(out_dir).resolve()
    out_root.mkdir(parents=True, exist_ok=True)

    all_stages = {"dart", "documents", "krx", "news", "ftc", "bok", "kosis"}
    only_set = {x.strip() for x in only.split(",") if x.strip()} if only else all_stages
    skip_set = {x.strip() for x in skip.split(",") if x.strip()}
    stages = only_set - skip_set

    logger.info(f"실행 단계: {sorted(stages)}")
    logger.info(f"기간: {from_year}~{to_year}")
    logger.info(f"산출물: {out_root}")

    needs_dart = bool(stages & {"dart", "documents"})
    if needs_dart and not DART_API_KEY:
        logger.error("DART_API_KEY가 .env에 없음 (dart/documents 단계 필수)")
        raise typer.Exit(1)
    if not DART_API_KEY:
        logger.warning("DART_API_KEY 없음 — dart/documents 단계는 건너뜀")

    cache_dir = out_root / "_common" / "dart_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    dart = DartHTTPClient(api_key=DART_API_KEY, cache_dir=cache_dir) if needs_dart else None

    catalog_path = Path(catalog).resolve()
    catalog_entries: list = []
    if catalog_path.exists():
        catalog_entries = load_catalog(catalog_path)
    elif needs_dart:
        logger.error(f"카탈로그 없음: {catalog_path}")
        raise typer.Exit(1)
    else:
        logger.warning(f"카탈로그 없음 (news 전용 실행이라 skip): {catalog_path}")
    logger.info(f"카탈로그 {len(catalog_entries)}항목 (callable={sum(1 for e in catalog_entries if e.callable_)})")
    write_json(out_root / "_common" / "catalog_snapshot.json",
               [asdict(e) for e in catalog_entries])
    # G2: catalog/8.다운로드_문서종류_선별에서 KEEP/DROP report_nm 자동 로드 (SSOT)
    load_report_filters_from_catalog(catalog_path)

    corp_cache = out_root / "_common" / "corp_code_cache.json"
    corps_list = normalize_corps(corps, corp_codes, corp_cache, dart)
    logger.info(f"대상 회사 {len(corps_list)}사: {[c['corp_name'] for c in corps_list]}")

    _started_iso = datetime.now().isoformat(timespec="seconds")
    manifest = Manifest(
        started=_started_iso,
        run_id=make_run_id(),
        hash_algo="sha1[:8]",  # see params_hash() — 파일명 접미 알고리즘 명시
        args={
            "corps": corps, "corp_codes": corp_codes,
            "from_year": from_year, "to_year": to_year,
            "stages": sorted(stages),
            "news_max_per_feed": news_max_per_feed,
            "news_since": news_since,
            "docs_max_per_corp": docs_max_per_corp,
        },
        corps=corps_list,
    )
    rcept_no_pool: dict[str, list[str]] = {}

    try:
        if "dart" in stages:
            logger.info("=== STAGE: DART API ===")
            rcept_no_pool = collect_dart(dart, corps_list, catalog_entries,
                                         out_root, manifest, from_year, to_year)
            write_json(out_root / "_common" / "rcept_no_pool.json", rcept_no_pool)
        else:
            pool_file = out_root / "_common" / "rcept_no_pool.json"
            if pool_file.exists():
                rcept_no_pool = json.loads(pool_file.read_text(encoding="utf-8"))

        if "documents" in stages:
            logger.info("=== STAGE: 공시 원문 ===")
            # G1: --no-rebuild-pool 플래그면 풀이 적어도 강제 유지 (PDF 재시도 등)
            need_rebuild = (not no_rebuild_pool) and (
                not rcept_no_pool or all(len(v) < 30 for v in rcept_no_pool.values())
            )
            if need_rebuild:
                logger.info("rcept_no_pool 재구성 (list 응답 파일 기반)")
                rebuilt = rebuild_pool_from_list_files(out_root, corps_list)
                if rebuilt:
                    rcept_no_pool = rebuilt
                    write_json(out_root / "_common" / "rcept_no_pool.json", rcept_no_pool)
                    logger.info("풀: " + ", ".join(f"{cc}={len(v)}" for cc, v in rcept_no_pool.items()))
            else:
                logger.info(f"풀 그대로 사용 ({sum(len(v) for v in rcept_no_pool.values())}개): "
                          + ", ".join(f"{cc}={len(v)}" for cc, v in rcept_no_pool.items()))
            collect_documents(dart, corps_list, rcept_no_pool, out_root, manifest, docs_max_per_corp)

        if "krx" in stages:
            logger.info("=== STAGE: KRX 일별 OHLCV (FinanceDataReader) ===")
            collect_krx_fdr(corps_list, out_root, manifest, from_year, to_year)

        if "news" in stages:
            logger.info("=== STAGE: 뉴스 RSS ===")
            since_d: date | None = None
            if news_since.strip():
                try:
                    since_d = date.fromisoformat(news_since.strip())
                except Exception:
                    logger.warning(f"news_since 파싱 실패 ({news_since}) — 필터 미적용")
            collect_news_wrap(corps_list, out_root, manifest, news_max_per_feed, since_d)

        if "ftc" in stages:
            logger.info("=== STAGE: 공정위 FTC ===")
            collect_ftc(out_root, manifest, corps_list, from_year, to_year,
                       catalog_entries=catalog_entries)

        if "bok" in stages:
            logger.info("=== STAGE: BOK ECOS ===")
            collect_bok(out_root, manifest, from_year, to_year)

        if "kosis" in stages:
            logger.info("=== STAGE: KOSIS ===")
            collect_kosis(out_root, manifest, from_year, to_year)
    finally:
        if dart is not None:
            dart.close()
        finalize_manifest(manifest, out_root)
        logger.info(f"manifest: {out_root / '_manifest.json'}")
        logger.info(f"요약: {manifest.summary}")


if __name__ == "__main__":
    app()
